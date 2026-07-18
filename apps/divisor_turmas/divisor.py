import streamlit as st
import pandas as pd
import datetime
import io
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Funções auxiliares de dados
# ---------------------------------------------------------------------------

def find_column(df, possible_names):
    """Localiza coluna no DataFrame ignorando maiúsculas/minúsculas."""
    for name in possible_names:
        for col in df.columns:
            if col.strip().lower() == name.lower():
                return col
    return None


def load_csv(file):
    """Carrega CSV com detecção automática de codificação e delimitador."""
    try:
        if isinstance(file, str):
            with open(file, 'rb') as f:
                sample = f.read(2048)
        else:
            sample = file.read(2048)
            file.seek(0)

        # Detectar codificação
        encoding_detected = 'utf-8'
        for enc in ['utf-8', 'iso-8859-1', 'cp1252']:
            try:
                sample.decode(enc)
                encoding_detected = enc
                break
            except UnicodeDecodeError:
                continue

        # Detectar delimitador
        sample_str = sample.decode(encoding_detected, errors='ignore')
        delimiter = ';'
        if ',' in sample_str and sample_str.count(',') > sample_str.count(';'):
            delimiter = ','

        if not isinstance(file, str):
            file.seek(0)

        df = pd.read_csv(file, sep=delimiter, encoding=encoding_detected)

        # Garante que a coluna "Identificação única" seja sempre tratada como texto
        for col in df.columns:
            if col.strip().lower() == 'identificação única':
                df[col] = df[col].astype(str)
                break

        return df
    except Exception as e:
        raise ValueError(f"Erro ao analisar o arquivo CSV: {str(e)}")


def load_xlsx(file):
    """Carrega XLSX do Educacenso normalizando para o mesmo layout do CSV.

    Suporta arquivos com/sem cabeçalho e rodapé institucional e com células
    mescladas: localiza a linha de cabeçalho pela âncora 'Ordem', mapeia os
    índices de coluna dos rótulos e extrai apenas as linhas de dados.
    """
    try:
        if not isinstance(file, str):
            file.seek(0)
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.worksheets[0]

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Planilha vazia.")

        def _norm(v):
            return str(v).strip().lower() if v is not None else ""

        # Localizar a linha de cabeçalho (âncora 'Ordem'; fallback 'Nome da turma').
        header_idx = None
        for i, row in enumerate(rows):
            normalized = [_norm(c) for c in row]
            if "ordem" in normalized or "nome da turma" in normalized:
                header_idx = i
                break
        if header_idx is None:
            raise ValueError("Linha de cabeçalho não encontrada (coluna 'Ordem').")

        header_row = rows[header_idx]
        # Índices de coluna com rótulo não vazio.
        col_map = [(idx, str(val).strip()) for idx, val in enumerate(header_row)
                   if val is not None and str(val).strip() != ""]
        col_indices = [idx for idx, _ in col_map]
        columns = [name for _, name in col_map]

        # Coluna âncora 'Ordem' para delimitar as linhas de dados.
        ordem_pos = None
        for pos, (idx, name) in enumerate(col_map):
            if name.strip().lower() == "ordem":
                ordem_pos = pos
                break

        data = []
        for row in rows[header_idx + 1:]:
            values = [row[idx] if idx < len(row) else None for idx in col_indices]
            # Parar no rodapé.
            first_txt = _norm(values[0])
            if first_txt.startswith("nota"):
                break
            # Manter apenas linhas cuja âncora 'Ordem' seja numérica.
            if ordem_pos is not None:
                anchor = values[ordem_pos]
                if not isinstance(anchor, (int, float)):
                    continue
            elif all(v is None or str(v).strip() == "" for v in values):
                continue
            data.append(values)

        wb.close()

        df = pd.DataFrame(data, columns=columns)
        df.columns = [str(c).strip() for c in df.columns]

        for col in df.columns:
            if col.strip().lower() == "identificação única":
                df[col] = df[col].astype(str)
                break

        return df
    except Exception as e:
        raise ValueError(f"Erro ao analisar o arquivo Excel: {str(e)}")


def load_data(file, filename):
    """Carrega o arquivo enviado escolhendo o leitor conforme a extensão."""
    name = str(filename).lower()
    if name.endswith(".xlsx"):
        return load_xlsx(file)
    return load_csv(file)


def _format_date_input():
    """Insere as barras (DD/MM/AAAA) automaticamente a partir dos dígitos."""
    raw = st.session_state.get("divisor_date_raw", "")
    digits = re.sub(r"\D", "", raw)[:8]
    if len(digits) >= 5:
        formatted = f"{digits[:2]}/{digits[2:4]}/{digits[4:]}"
    elif len(digits) >= 3:
        formatted = f"{digits[:2]}/{digits[2:]}"
    else:
        formatted = digits
    st.session_state["divisor_date_raw"] = formatted


def sanitize_filename(name):
    """Sanitiza string para uso seguro em nomes de arquivo."""
    s = str(name).strip().replace(" ", "_")
    s = re.sub(r'(?u)[^-\w.]', '', s)
    return s


def sanitize_sheet_title(name):
    """Sanitiza nome de aba respeitando limite do Excel (31 chars)."""
    sanitized = re.sub(r'[\\/*?:\[\]]', '', str(name))
    return sanitized[:31]


def generate_xlsx(df, school_name, collection_date):
    """Gera o workbook Excel formatado e retorna buffer binário."""
    class_col = find_column(df, ["Nome da turma", "turma", "nome turma", "código da turma"])
    name_col  = find_column(df, ["Nome", "nome do aluno", "aluno", "nome"])

    if not class_col:
        raise ValueError("Coluna 'Nome da turma' não encontrada.")
    if not name_col:
        raise ValueError("Coluna 'Nome' não encontrada.")

    df_clean = df.copy()
    df_clean[class_col] = df_clean[class_col].fillna("SEM_TURMA")
    classes = df_clean[class_col].unique()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    font_info   = Font(name="Calibri", size=11, bold=True)
    font_header = Font(name="Calibri", size=11, bold=True)
    font_data   = Font(name="Calibri", size=11)
    fill_header = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    align_left  = Alignment(horizontal="left", vertical="center")

    for class_name in sorted(classes):
        ws = wb.create_sheet(title=sanitize_sheet_title(class_name))

        df_class = df_clean[df_clean[class_col] == class_name]
        df_class = df_class.sort_values(by=name_col, ascending=True)

        # 3 linhas de cabeçalho
        ws.cell(row=1, column=1, value=f"Unidade Escolar: {school_name}")
        ws.cell(row=2, column=1, value=f"Data da coleta: {collection_date}")
        ws.cell(row=3, column=1, value=f"Turma: {class_name}")
        for r in range(1, 4):
            ws.cell(row=r, column=1).font      = font_info
            ws.cell(row=r, column=1).alignment = align_left

        # Linha 4 — cabeçalho da tabela
        for col_idx, header_val in enumerate(df_class.columns, 1):
            cell = ws.cell(row=4, column=col_idx, value=header_val)
            cell.font      = font_header
            cell.fill      = fill_header
            cell.alignment = align_left

        # Dados a partir da linha 5 com renumeração da coluna 1
        for row_idx, row_values in enumerate(df_class.values, 5):
            for col_idx, cell_value in enumerate(row_values, 1):
                val = "" if pd.isna(cell_value) else cell_value
                if col_idx == 1:
                    val = row_idx - 4   # renumeração sequencial: 1, 2, 3…
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font      = font_data
                cell.alignment = align_left

        # Largura automática das colunas
        for col in ws.columns:
            max_len    = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col[3:]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(min(max_len + 3, 50), 10)

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer


# ---------------------------------------------------------------------------
# Página do DIVISOR DE ALUNOS POR TURMAS
# ---------------------------------------------------------------------------

def show_divisor_page(logo_base64):
    """Renderiza a interface completa do DIVISOR DE ALUNOS POR TURMAS."""

    # ── Banner ──────────────────────────────────────────────────────────────
    if logo_base64:
        banner_html = f"""
        <div class="title-container" style="display:flex; align-items:center; gap:1.5rem;">
            <img src="data:image/png;base64,{logo_base64}"
                 style="width:68px; height:68px; border-radius:12px;
                        box-shadow:0 4px 12px rgba(0,0,0,0.2);
                        background:#fff; padding:5px;" />
            <div>
                <h1 style="margin:0; padding:0; line-height:1.2; color:white !important;">
                    Relatórios de Alunos (Organizar por Turma)
                </h1>
                <p style="margin-top:0.3rem; color:#bfdbfe; font-size:1rem; margin-bottom:0;">
                    Organize o Relatório de Alunos do Educacenso (CSV ou XLSX) em abas separadas por turma.
                </p>
            </div>
        </div>"""
    else:
        banner_html = """
        <div class="title-container">
            <h1>🏫 Relatórios de Alunos (Organizar por Turma)</h1>
            <p>Organize as planilhas do Educacenso em abas separadas por turma.</p>
        </div>"""

    st.markdown(banner_html, unsafe_allow_html=True)

    # ── Layout de entrada ────────────────────────────────────────────────────
    col_inputs, col_preview = st.columns([1, 1])

    with col_inputs:
        st.markdown('<div class="custom-card">', unsafe_allow_html=True)
        st.subheader("Dados de Identificação")

        school_name = st.text_input(
            "Nome da Unidade Escolar",
            placeholder="Digite o nome completo da escola...",
            help="Será exibido na primeira linha de todas as abas."
        )

        collection_date_str = st.text_input(
            "Data do download do arquivo (DD/MM/AAAA)",
            placeholder="Ex: 06072026",
            help="Digite apenas os números (DDMMAAAA). As barras são inseridas automaticamente.",
            key="divisor_date_raw",
            on_change=_format_date_input,
        )

        is_date_valid = False
        if collection_date_str:
            if re.match(r"^\d{2}/\d{2}/\d{4}$", collection_date_str.strip()):
                try:
                    d, m, y = [int(x) for x in collection_date_str.strip().split('/')]
                    datetime.date(y, m, d)
                    is_date_valid = True
                except ValueError:
                    st.error("⚠️ Data inválida. Verifique o dia, mês e ano.")
            else:
                st.error("⚠️ Formato incorreto. Use DD/MM/AAAA (ex: 06/07/2026).")

                
        st.subheader("📥 Enviar Arquivo (CSV ou Excel)")
        uploaded_file = st.file_uploader(
            "Arquivo do Educacenso (CSV ou XLSX)",
            type=["csv", "xlsx"],
            help="Arquivo baixado do Educacenso em formato CSV ou XLSX."
        )
        st.markdown('</div>', unsafe_allow_html=True)

    # ── Processamento ────────────────────────────────────────────────────────
    df          = None
    file_loaded = False

    if uploaded_file is not None:
        try:
            df          = load_data(uploaded_file, uploaded_file.name)
            file_loaded = True
        except Exception as e:
            st.error(f"Erro ao processar o arquivo: {e}")

    if file_loaded and df is not None:
        with col_preview:
            st.markdown('<div class="custom-card">', unsafe_allow_html=True)
            st.subheader("📊 Resumo dos Dados")

            class_col = find_column(df, ["Nome da turma", "turma", "nome turma", "código da turma"])
            name_col  = find_column(df, ["Nome", "nome do aluno", "aluno", "nome"])

            if class_col and name_col:
                unique_classes = df[class_col].dropna().unique()
                st.markdown(f"""
                <table>
                    <thead>
                        <tr><th>Indicador</th><th style="text-align:right;">Total</th></tr>
                    </thead>
                    <tbody>
                        <tr><td>Total de Alunos</td><td style="text-align:right; font-weight:700; color:var(--text-primary);">{len(df)}</td></tr>
                        <tr><td>Total de Turmas</td><td style="text-align:right; font-weight:700; color:var(--text-primary);">{len(unique_classes)}</td></tr>
                    </tbody>
                </table>
                """, unsafe_allow_html=True)
            else:
                st.warning("⚠️ Colunas 'Nome da turma' ou 'Nome' não encontradas.")
                st.info("Colunas disponíveis: " + ", ".join(df.columns))

            st.markdown('</div>', unsafe_allow_html=True)

        # ── Ações / Download ─────────────────────────────────────────────────
        st.markdown('<div class="custom-card">', unsafe_allow_html=True)
        st.subheader("⚙️ Ações e Download")

        if not school_name:
            st.warning("⚠️ Informe o Nome da Unidade Escolar para continuar.")
        elif not collection_date_str:
            st.warning("⚠️ Informe a Data do download do arquivo para continuar.")
        elif not is_date_valid:
            st.error("⚠️ Data inválida. Corrija antes de gerar o arquivo.")
        else:
            try:
                with st.spinner("Processando dados e organizando abas..."):
                    excel_data = generate_xlsx(df, school_name, collection_date_str.strip())

                today_str        = datetime.date.today().strftime("%d_%m_%Y")
                output_filename  = f"{sanitize_filename(school_name)}_{today_str}.xlsx"

                st.success("✅ Arquivo gerado com sucesso!")

                st.download_button(
                    label="📥 Baixar Excel (.xlsx)",
                    data=excel_data,
                    file_name=output_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            except Exception as e:
                st.error(f"Erro ao gerar planilha: {e}")

        st.markdown('</div>', unsafe_allow_html=True)

    else:
        with col_preview:
            st.markdown('<div class="custom-card">', unsafe_allow_html=True)
            st.subheader("ℹ️ Como funciona?")
            st.markdown("""
            1. **Preencha** o nome da escola e a data do downlaod do arquivo.
            2. **Envie** o Relatório de Alunos baixado do Educacenso, em formato CSV ou XLSX.
            3. O sistema irá:
               - Agrupar alunos por turma (**Nome da turma**).
               - Ordenar em ordem alfabética crescente (**Nome**).
               - Gerar cabeçalho customizado de 3 linhas por aba.
               - Renumerar a primeira coluna sequencialmente (1, 2, 3…).
               - Ajustar **automaticamente** a largura das colunas.
            4. **Baixe** o arquivo Excel formatado.
            """)
            st.markdown('</div>', unsafe_allow_html=True)
